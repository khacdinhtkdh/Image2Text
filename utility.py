from openpyxl import Workbook as WB

STT = 'STT'
FULL_NAME = 'FULL NAME'
BIRTHDAY = 'BIRTHDAY'
PHONE = 'PHONE'
EMAIL = 'EMAIL'
PASS_PAYPAL = 'PASS PAYPAL'
PASS_EMAIL = 'PASS EMAIL'
MAIL_KHOI_PHUC = 'MAIL KHOI PHUC'
ADDRESS = 'ADDRESS'
ZIPCODE = 'ZIPCODE'
ID = 'ID'
CARD = 'CARD'
NGAY_REG = 'NGAY REG'
NOTE = 'NOTE'
VISA = 'VISA'
CHUYEN_DANG_FRIEND = 'CHUYEN DANG FRIEND'
CHUYEN_DANG_GOOD = 'CHUYEN DANG GOOD'
FOLDER = 'FOLDER'


def print_excel(data, savepath):
    # Export data to Excel file
    output = WB()
    sheet = output.active
    sheet.cell(row=1, column=1).value = STT
    sheet.cell(row=1, column=2).value = FULL_NAME
    sheet.cell(row=1, column=3).value = BIRTHDAY
    sheet.cell(row=1, column=4).value = PHONE
    sheet.cell(row=1, column=5).value = EMAIL
    sheet.cell(row=1, column=6).value = PASS_PAYPAL
    sheet.cell(row=1, column=7).value = PASS_EMAIL
    sheet.cell(row=1, column=8).value = MAIL_KHOI_PHUC
    sheet.cell(row=1, column=9).value = ADDRESS
    sheet.cell(row=1, column=10).value = ZIPCODE
    sheet.cell(row=1, column=11).value = ID
    sheet.cell(row=1, column=12).value = CARD
    sheet.cell(row=1, column=13).value = NGAY_REG
    sheet.cell(row=1, column=14).value = NOTE
    sheet.cell(row=1, column=15).value = CHUYEN_DANG_FRIEND
    sheet.cell(row=1, column=16).value = CHUYEN_DANG_GOOD
    sheet.cell(row=1, column=17).value = FOLDER

    for i in range(len(data)):
        sheet.cell(row=i + 2, column=1).value = i + 1
        if data[i].get(FULL_NAME):
            sheet.cell(row=i + 2, column=2).value = data[i][FULL_NAME]
        if data[i].get(BIRTHDAY):
            sheet.cell(row=i + 2, column=3).value = data[i][BIRTHDAY]
        if data[i].get(PASS_PAYPAL):
            sheet.cell(row=i + 2, column=6).value = data[i][PASS_PAYPAL]
        if data[i].get(ADDRESS):
            sheet.cell(row=i + 2, column=9).value = data[i][ADDRESS]
        if data[i].get(ZIPCODE):
            sheet.cell(row=i + 2, column=10).value = data[i][ZIPCODE]
        if data[i].get(ID):
            sheet.cell(row=i + 2, column=11).value = data[i][ID]
        if data[i].get(FOLDER):
            sheet.cell(row=i + 2, column=17).value = data[i][FOLDER]
    output.save(savepath)
